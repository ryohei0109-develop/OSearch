"""Excel ファイルのパーサー（セル＋オートシェイプ）"""

import zipfile
from lxml import etree
import openpyxl


# XML 名前空間
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"


def parse(file_path: str) -> list[dict]:
    """
    .xlsx ファイルを解析してコンテンツリストを返す。
    セル内容とオートシェイプ内テキストの両方を対象とする。
    """
    contents = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"[警告] Excelファイルの読み込みに失敗しました: {file_path} ({e})")
        return contents

    # --- セル内容 ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                contents.append({
                    "location": {
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "slide": None,
                        "page": None,
                        "shape_id": None,
                    },
                    "content": text,
                })

    # --- オートシェイプ内テキスト（XML解析）---
    try:
        contents.extend(_parse_drawings(file_path))
    except Exception as e:
        print(f"[警告] Excelオートシェイプ解析に失敗しました: {file_path} ({e})")

    return contents


def _parse_drawings(file_path: str) -> list[dict]:
    """xl/drawings/*.xml から a:t タグのテキストを抽出する。"""
    results = []
    with zipfile.ZipFile(file_path, "r") as zf:
        drawing_files = [n for n in zf.namelist() if n.startswith("xl/drawings/") and n.endswith(".xml")]
        for drawing_path in drawing_files:
            drawing_id = drawing_path.split("/")[-1].replace(".xml", "")
            try:
                xml_data = zf.read(drawing_path)
                root = etree.fromstring(xml_data)
                for t_elem in root.iter(f"{{{_NS_A}}}t"):
                    text = (t_elem.text or "").strip()
                    if not text:
                        continue
                    results.append({
                        "location": {
                            "sheet": None,
                            "cell": None,
                            "slide": None,
                            "page": None,
                            "shape_id": drawing_id,
                        },
                        "content": text,
                    })
            except Exception as e:
                print(f"[警告] Drawing XML の解析に失敗しました: {drawing_path} ({e})")

    return results
